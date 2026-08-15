from __future__ import annotations

import csv
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from chdd_model import DEFAULT_ASSUMPTIONS, DEFAULT_PUMPS, METHODOLOGY_LOCKS, REQUIRED_COLUMNS


RUSSIAN_INPUT_HEADERS = [
    "Дата",
    "Скважина",
    "Накопленная добыча жидкости, т",
    "Дебит жидкости м3/сут.",
    "Накопленная добыча нефти, т.",
    "Дебит нефти, т./сут.",
    "Приемистость, м3/сут.",
    "Накопленная объем закачки, м3",
    "Пластовое давление",
    "Забойное давление",
    "Коэффициент эксплуатации",
    "Добыча жидкости за месяц, т.",
    "Добыча нефти за месяц, т.",
    "Закачка воды за месяц, м3",
]


def _normalized_header(value: Any) -> str:
    text = str(value or "").strip().casefold().replace("ё", "е").replace("³", "3")
    return re.sub(r"[^a-zа-я0-9]+", "", text)


HEADER_ALIASES: dict[str, str] = {}
for canonical in REQUIRED_COLUMNS:
    HEADER_ALIASES[_normalized_header(canonical)] = canonical
for russian, canonical in zip(RUSSIAN_INPUT_HEADERS, REQUIRED_COLUMNS):
    HEADER_ALIASES[_normalized_header(russian)] = canonical

for alias, canonical in {
    "Накопленный объем закачки, м3": "WWIT",
    "Накопленный объем закачки воды, м3": "WWIT",
    "Добыча жидкости за месяц, т": "WLPT_Diff",
    "Добыча нефти за месяц, т": "WOMT_Diff",
    "Закачка за месяц, м3": "WWIT_Diff",
    "Номер скважины": "well",
}.items():
    HEADER_ALIASES[_normalized_header(alias)] = canonical


def _canonicalize_headers(raw_headers: Iterable[Any]) -> list[str]:
    headers: list[str] = []
    seen: set[str] = set()
    for raw in raw_headers:
        canonical = HEADER_ALIASES.get(_normalized_header(raw), str(raw or "").strip())
        if canonical in seen:
            raise ValueError(f"После сопоставления заголовков найден дублирующийся столбец: {canonical}")
        seen.add(canonical)
        headers.append(canonical)
    return headers


def load_source_records(path: str | Path) -> tuple[list[str], list[dict[str, Any]]]:
    source_path = Path(path)
    if not source_path.exists():
        raise FileNotFoundError(f"Файл исходных данных не найден: {source_path}")
    suffix = source_path.suffix.casefold()

    if suffix == ".xlsx":
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        sheet = workbook["MODEL_Y"] if "MODEL_Y" in workbook.sheetnames else workbook.active
        values = sheet.iter_rows(values_only=True)
        try:
            raw_headers = list(next(values))
        except StopIteration as exc:
            raise ValueError("Исходный Excel-файл пуст.") from exc
        headers = _canonicalize_headers(raw_headers)
        records = [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in values
            if any(value not in (None, "") for value in row)
        ]
        workbook.close()
        return headers, records

    if suffix == ".csv":
        text = source_path.read_text(encoding="utf-8-sig")
        try:
            dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","
        rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
        if not rows:
            raise ValueError("Исходный CSV-файл пуст.")
        headers = _canonicalize_headers(rows[0])
        records = [
            {headers[index]: value if index < len(row) else "" for index, value in enumerate(row[: len(headers)])}
            for row in rows[1:]
            if any(str(value).strip() for value in row)
        ]
        return headers, records

    raise ValueError("Поддерживаются исходные файлы .xlsx и .csv.")


def load_norms(path: str | Path) -> tuple[dict[str, Any], list[dict[str, float]]]:
    norms_path = Path(path)
    if not norms_path.exists():
        raise FileNotFoundError(f"Файл нормативов не найден: {norms_path}")
    workbook = load_workbook(norms_path, read_only=True, data_only=True)
    if "Нормативы" not in workbook.sheetnames or "ЭЦН" not in workbook.sheetnames:
        raise ValueError("В файле нормативов должны быть листы «Нормативы» и «ЭЦН».")

    assumptions = dict(DEFAULT_ASSUMPTIONS)
    for row in workbook["Нормативы"].iter_rows(min_row=2, values_only=True):
        code, _, value = (list(row) + [None, None, None])[:3]
        if not code or code not in assumptions:
            continue
        if isinstance(assumptions[code], bool):
            assumptions[code] = str(value).strip().casefold() in {"1", "true", "да", "истина"}
        else:
            try:
                assumptions[code] = float(value)
            except (TypeError, ValueError):
                raise ValueError(f"Норматив {code} должен быть числом, получено: {value!r}") from None

    pumps: list[dict[str, float]] = []
    for row in workbook["ЭЦН"].iter_rows(min_row=2, values_only=True):
        nominal, minimum, maximum, cost = (list(row) + [None, None, None, None])[:4]
        if nominal in (None, ""):
            continue
        try:
            pumps.append(
                {
                    "nominal": float(nominal),
                    "min": float(minimum),
                    "max": float(maximum),
                    "costM": float(cost),
                }
            )
        except (TypeError, ValueError):
            raise ValueError(f"Некорректная строка таблицы ЭЦН: {row}") from None
    workbook.close()
    if not pumps:
        raise ValueError("В файле нормативов отсутствует таблица ЭЦН.")
    return assumptions, pumps


def save_result_json(result: dict[str, Any], path: str | Path) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path


def create_norms_workbook(path: str | Path) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Нормативы"
    sheet.sheet_view.showGridLines = False
    headers = ["Код", "Параметр", "Значение", "Ед. изм.", "Можно менять", "Комментарий методики"]
    sheet.append(headers)

    definitions = [
        ("oilPriceRubT", "Приведенная цена реализации нефти", 28_000, "руб./т", True, "Единый норматив на весь горизонт."),
        ("deductionsRubT", "Отчисления на тонну нефти", 19_600, "руб./т", True, "Вычитаются при расчете EBITDA и налоговой базы."),
        ("oilOpexRubT", "OPEX на добычу нефти", 40, "руб./т", True, "Переменные затраты по нефти."),
        ("liquidOpexRubT", "OPEX на добычу жидкости", 100, "руб./т", True, "Переменные затраты по жидкости."),
        ("injectionOpexRubM3", "OPEX на закачку воды", 30, "руб./м³", True, "Переменные затраты по закачке."),
        ("fundAnnualRubWell", "Содержание действующего фонда", 1_000_000, "руб./скв.-год", True, "Начисляется за каждый активный скважино-месяц."),
        ("pumpOperationCostM", "Операция по смене ЭЦН", 1.8, "млн руб./операцию", True, "OPEX, отдельно от CAPEX нового насоса."),
        ("stopStartCostM", "Остановка или запуск скважины", 1.0, "млн руб./событие", False, "Зафиксировано методикой: остановка и последующий запуск — два платных события."),
        ("conversionBaseCostM", "Перевод добывающей скважины под закачку", 5.0, "млн руб./перевод", False, "Зафиксировано методикой; без отдельной платы за остановку и без старого ЭЦН."),
        ("profitTaxRate", "Ставка налога на прибыль", 25.0, "%", True, "Налог = max(0; годовая налоговая база) × ставка."),
        ("propertyTaxRate", "Ставка налога на имущество", 2.2, "%", True, "От средней остаточной стоимости."),
        ("waccRate", "Ставка дисконтирования WACC", 10.0, "%", True, "Для стартового года DF = 1."),
        ("annualDepreciationM", "Годовая амортизация", 0.0, "млн руб./год", True, "Распределяется по имеющимся месяцам как 1/12."),
        ("residualStartM", "Остаточная стоимость на начало года", 0.0, "млн руб.", True, "Единая величина по годам, если задана."),
        ("residualEndM", "Остаточная стоимость на конец года", 0.0, "млн руб.", True, "Единая величина по годам, если задана."),
        ("otherIncludedEbitdaM", "Прочие ДиР, включаемые в EBITDA", 0.0, "млн руб./год", True, "Вычитаются из EBITDA и налоговой базы."),
        ("otherExcludedEbitdaM", "Прочие ДиР, не включаемые в EBITDA", 0.0, "млн руб./год", True, "Вычитаются только из налоговой базы."),
        ("chargeInitialPump", "Начислять начальный ЭЦН", "НЕТ", "да/нет", True, "По умолчанию существующий на старте ЭЦН не оплачивается."),
        ("downsizeThreshold", "Порог уменьшения типоразмера ЭЦН", 100.0, "м³/сут", False, "Зафиксировано методикой: замена вниз только при превышении 100 м³/сут."),
    ]
    for code, label, value, unit, can_change, comment in definitions:
        sheet.append((code, label, value, unit, "ДА" if can_change else "НЕТ", comment))

    pump_sheet = workbook.create_sheet("ЭЦН")
    pump_sheet.sheet_view.showGridLines = False
    pump_sheet.append(["Номинал, м³/сут", "Минимум, м³/сут", "Максимум, м³/сут", "Стоимость нового ЭЦН, млн руб."])
    for pump in DEFAULT_PUMPS:
        pump_sheet.append([pump["nominal"], pump["min"], pump["max"], pump["costM"]])

    dark = PatternFill("solid", fgColor="17324D")
    light_blue = PatternFill("solid", fgColor="DDEBF7")
    locked_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="CAD6E2")
    for current_sheet in (sheet, pump_sheet):
        for cell in current_sheet[1]:
            cell.fill = dark
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="medium", color="17324D"))
        current_sheet.freeze_panes = "A2"
        current_sheet.auto_filter.ref = current_sheet.dimensions
        current_sheet.row_dimensions[1].height = 36
        for row in current_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_index in range(2, sheet.max_row + 1):
        can_change = str(sheet.cell(row_index, 5).value).strip().casefold() == "да"
        sheet.cell(row_index, 3).font = Font(color="0070C0", bold=True)
        sheet.cell(row_index, 3).fill = light_blue if can_change else locked_fill
        if not can_change:
            sheet.cell(row_index, 5).font = Font(color="C00000", bold=True)
    yes_no = DataValidation(type="list", formula1='"ДА,НЕТ"', allow_blank=False)
    sheet.add_data_validation(yes_no)
    charge_row = next(index for index in range(2, sheet.max_row + 1) if sheet.cell(index, 1).value == "chargeInitialPump")
    yes_no.add(sheet.cell(charge_row, 3))

    for row_index in range(2, pump_sheet.max_row + 1):
        for column in range(1, 5):
            pump_sheet.cell(row_index, column).font = Font(color="0070C0")
        pump_sheet.cell(row_index, 4).number_format = "0.00000"

    widths = [27, 43, 18, 20, 16, 64]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for column, width in enumerate([21, 21, 23, 34], 1):
        pump_sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.orientation = "landscape"
    pump_sheet.sheet_properties.pageSetUpPr.fitToPage = True
    pump_sheet.page_setup.fitToWidth = 1
    pump_sheet.page_setup.fitToHeight = 1
    pump_sheet.page_setup.orientation = "landscape"
    workbook.save(output_path)
    return output_path


def create_input_template(path: str | Path, *, with_example: bool = True) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MODEL_Y"
    sheet.sheet_view.showGridLines = False
    sheet.append(RUSSIAN_INPUT_HEADERS)

    if with_example:
        rows: list[list[Any]] = []
        cumulative = {
            "A": {"liquid": 0.0, "oil": 0.0, "injection": 0.0},
            "B": {"liquid": 0.0, "oil": 0.0, "injection": 0.0},
        }
        for year, month in [(2013, 12)] + [(year, month) for year in (2014, 2015) for month in range(1, 13)]:
            current_date = date(year, month, 1)
            if year == 2013:
                a_liq, a_oil, a_wlpr, a_womr = 0.0, 0.0, 100.0, 10.0
            elif year == 2014 and month == 2:
                a_liq, a_oil, a_wlpr, a_womr = 0.0, 0.0, 0.0, 0.0
            elif year == 2014 and month >= 4:
                a_liq, a_oil, a_wlpr, a_womr = 4_200.0, 900.0, 170.0, 36.0
            elif year == 2015:
                a_liq, a_oil, a_wlpr, a_womr = 650.0, 180.0, 20.0, 6.0
            else:
                a_liq, a_oil, a_wlpr, a_womr = 2_600.0, 700.0, 100.0, 24.0
            cumulative["A"]["liquid"] += a_liq
            cumulative["A"]["oil"] += a_oil
            rows.append([
                current_date,
                "A",
                cumulative["A"]["liquid"],
                a_wlpr,
                cumulative["A"]["oil"],
                a_womr,
                0.0,
                cumulative["A"]["injection"],
                105.0,
                80.0,
                1.0,
                a_liq,
                a_oil,
                0.0,
            ])

            if year == 2013 or (year == 2014 and month == 1):
                b_liq, b_oil, b_wlpr, b_womr, b_wwir, b_inj = 1_900.0, 500.0, 80.0, 18.0, 0.0, 0.0
            elif year == 2015 and month == 1:
                b_liq, b_oil, b_wlpr, b_womr, b_wwir, b_inj = 0.0, 0.0, 0.0, 0.0, 0.0, 0.0
            else:
                b_liq, b_oil, b_wlpr, b_womr, b_wwir, b_inj = 0.0, 0.0, 0.0, 0.0, 80.0, 2_200.0
            cumulative["B"]["liquid"] += b_liq
            cumulative["B"]["oil"] += b_oil
            cumulative["B"]["injection"] += b_inj
            rows.append([
                current_date,
                "B",
                cumulative["B"]["liquid"],
                b_wlpr,
                cumulative["B"]["oil"],
                b_womr,
                b_wwir,
                cumulative["B"]["injection"],
                102.0,
                76.0,
                1.0,
                b_liq,
                b_oil,
                b_inj,
            ])
        for row in rows:
            sheet.append(row)

    dark = PatternFill("solid", fgColor="17324D")
    thin = Side(style="thin", color="D9E2F3")
    for cell in sheet[1]:
        cell.fill = dark
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=thin)
        row[0].number_format = "dd.mm.yyyy"
    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 48
    for index, width in enumerate([13, 14, 24, 21, 23, 20, 20, 25, 20, 20, 23, 24, 23, 23], 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.orientation = "landscape"
    workbook.save(output_path)
    return output_path


def write_annual_report(result: dict[str, Any], path: str | Path) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ЧДД по годам"
    sheet.sheet_view.showGridLines = False

    years = [item["year"] for item in result["annual"]]
    first_year_column = 4
    last_column = first_year_column + len(years) - 1
    last_letter = get_column_letter(last_column)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    sheet["A1"] = "РАСЧЕТ ЧИСТОГО ДИСКОНТИРОВАННОГО ДОХОДА (ЧДД)"
    sheet["A1"].fill = PatternFill("solid", fgColor="17324D")
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 30

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    sheet["A2"] = (
        f"Источник: {result.get('sourceFile') or 'не указан'}  |  "
        f"Период: {result['startDate'][:4]}–{result['maxDate'][:4]}  |  "
        f"Версия ядра: {result['version']}"
    )
    sheet["A2"].font = Font(color="52677B", italic=True, size=9)

    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_column)
    sheet["A3"] = (
        f"Итоговый ЧДД: {result['summary']['totalChddM']:,.3f} млн руб.   |   "
        f"ИДДз: {result['summary']['profitabilityIndex']:,.3f}   |   "
        f"Остановки/запуски: {result['summary']['startStopCount']:.0f} событий   |   "
        f"Переводы под закачку: {result['summary']['conversionCount']:.0f}"
    ).replace(",", " ")
    sheet["A3"].fill = PatternFill("solid", fgColor="DDEBF7")
    sheet["A3"].font = Font(color="17324D", bold=True, size=11)
    sheet["A3"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[3].height = 24

    header_row = 5
    headers = ["Показатель", "Ед. изм.", "Методика / источник"] + years
    for column, value in enumerate(headers, 1):
        cell = sheet.cell(header_row, column, value)
        cell.fill = PatternFill("solid", fgColor="284B63")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 34

    metric_sections: list[tuple[str, str, list[tuple[str, str, str, str, str]]]] = [
        (
            "Исходные производственные показатели",
            "5B9BD5",
            [
                ("sourceCumulativeLiquidKt", "Накопленная добыча жидкости", "тыс. т", "Сумма последних счетчиков WLPT по скважинам на конец года", "number"),
                ("sourceCumulativeOilKt", "Накопленная добыча нефти", "тыс. т", "Сумма последних счетчиков WOMT по скважинам на конец года", "number"),
                ("sourceCumulativeInjectionKm3", "Накопленная закачка воды", "тыс. м³", "Сумма последних счетчиков WWIT по скважинам на конец года", "number"),
                ("liquidKt", "Добыча жидкости за год", "тыс. т", "Σ WLPT_Diff / 1 000", "number"),
                ("oilKt", "Добыча нефти за год", "тыс. т", "Σ WOMT_Diff / 1 000", "number"),
                ("injectionKm3", "Закачка воды за год", "тыс. м³", "Σ WWIT_Diff / 1 000", "number"),
                ("averageActiveWells", "Средний действующий фонд", "скв.", "Активна при WLPR > 0 или WWIR > 0", "number"),
                ("activeWellMonths", "Активный фонд", "скв.-мес.", "Σ активных скважин по месяцам", "integer"),
                ("averageWLPR", "Средний дебит жидкости", "м³/сут", "Среднее по положительным WLPR", "number"),
                ("averageWOMR", "Средний дебит нефти", "т/сут", "Среднее по положительным WOMR", "number"),
                ("averageWWIR", "Средняя приемистость", "м³/сут", "Среднее по положительным WWIR", "number"),
            ],
        ),
        (
            "Действия с фондом и платные события",
            "ED7D31",
            [
                ("pumpChanges", "Установлено новых ЭЦН — всего", "шт.", "Платные замены ЭЦН; начальный насос не начисляется по умолчанию", "integer"),
                ("pumpUps", "Смен ЭЦН на больший типоразмер", "шт.", "При выходе дебита выше максимума текущего ЭЦН", "integer"),
                ("pumpDowns", "Смен ЭЦН на меньший типоразмер", "шт.", "Только если min текущего ЭЦН − WLPR > 100", "integer"),
                ("stopCount", "Остановки скважин", "шт.", "Каждый переход активна → неактивна", "integer"),
                ("startCount", "Запуски скважин", "шт.", "Каждый переход неактивна → активна", "integer"),
                ("startStopCount", "Остановки и запуски — всего", "шт.", "Остановка и последующий запуск считаются двумя событиями", "integer"),
                ("conversionCount", "Переводы под закачку", "шт.", "Прямой переход добывающая → нагнетательная", "integer"),
                ("pumpOperationM", "Стоимость операций по смене ЭЦН", "млн руб.", "1,8 млн руб. за каждую замену", "money"),
                ("startStopCostM", "Стоимость остановок и запусков", "млн руб.", "1,0 млн руб. за каждое событие", "money"),
                ("conversionOpexM", "Стоимость переводов под закачку", "млн руб.", "5,0 млн руб. за перевод; без старого ЭЦН и отдельной остановки", "money"),
            ],
        ),
        (
            "Исходные экономические нормативы",
            "70AD47",
            [
                ("@oilPriceRubT", "Приведенная цена реализации нефти", "руб./т", "Файл нормативов", "rubles"),
                ("@deductionsRubT", "Отчисления", "руб./т", "Файл нормативов", "rubles"),
                ("@oilOpexRubT", "OPEX на нефть", "руб./т", "Файл нормативов", "rubles"),
                ("@liquidOpexRubT", "OPEX на жидкость", "руб./т", "Файл нормативов", "rubles"),
                ("@injectionOpexRubM3", "OPEX на закачку", "руб./м³", "Файл нормативов", "rubles"),
                ("@fundAnnualRubWell", "Содержание действующего фонда", "руб./скв.-год", "Файл нормативов", "rubles"),
                ("@profitTaxRate", "Ставка налога на прибыль", "%", "Файл нормативов", "percent"),
                ("@propertyTaxRate", "Ставка налога на имущество", "%", "Файл нормативов", "percent"),
                ("@waccRate", "WACC", "%", "Файл нормативов", "percent"),
            ],
        ),
        (
            "Рассчитанные экономические показатели",
            "4472C4",
            [
                ("revenueM", "Выручка от основной деятельности", "млн руб.", "Добыча нефти × цена", "money"),
                ("oilOpexM", "OPEX на нефть", "млн руб.", "Добыча нефти × норматив", "money"),
                ("liquidOpexM", "OPEX на жидкость", "млн руб.", "Добыча жидкости × норматив", "money"),
                ("injectionOpexM", "OPEX на закачку", "млн руб.", "Закачка × норматив", "money"),
                ("fundOpexM", "Содержание действующего фонда", "млн руб.", "Активные скважино-месяцы × годовой норматив / 12", "money"),
                ("propertyTaxM", "Налог на имущество", "млн руб.", "((ОС нач. + ОС кон.) / 2) × 2,2%", "money"),
                ("opexM", "Полный OPEX", "млн руб.", "Все операционные затраты, включая события", "money"),
                ("capexM", "CAPEX новых ЭЦН", "млн руб.", "Стоимость установленных новых насосов", "money"),
                ("depreciationM", "Амортизация", "млн руб.", "Годовой норматив / 12 по имеющимся месяцам", "money"),
                ("deductionsM", "Отчисления", "млн руб.", "Добыча нефти × норматив", "money"),
                ("taxableProfitM", "Прибыль до налогообложения", "млн руб.", "Выручка − OPEX − амортизация − отчисления − ДиР", "money"),
                ("profitTaxM", "Налог на прибыль", "млн руб.", "max(0; годовая налоговая база) × 25%", "money"),
                ("ebitdaM", "EBITDA", "млн руб.", "Выручка − OPEX − отчисления − ДиР в EBITDA", "money"),
                ("fcfM", "FCF", "млн руб.", "EBITDA − налог на прибыль − CAPEX", "money"),
                ("discountFactor", "Коэффициент дисконтирования", "доля", "1 / (1 + WACC)^(год − стартовый год)", "ratio"),
                ("chddM", "ЧДД", "млн руб.", "FCF × коэффициент дисконтирования", "money"),
                ("cumulativeChddM", "Накопленный ЧДД", "млн руб.", "Накопительная сумма ЧДД", "money"),
                ("discountedInflowM", "Дисконтированные притоки", "млн руб.", "Для расчета ИДДз", "money"),
                ("discountedOutflowM", "Дисконтированные оттоки", "млн руб.", "CAPEX + OPEX + отчисления + налог + положительные ДиР", "money"),
                ("profitabilityIndex", "ИДДз", "д. ед.", "Накопленные дисконтированные притоки / оттоки", "ratio"),
            ],
        ),
    ]

    pump_rows = [
        (
            f"#pump:{pump['nominal']:g}",
            f"ЭЦН {pump['nominal']:g} м³/сут — установлено новых",
            "шт.",
            "Количество оплаченных новых насосов данного типоразмера",
            "integer",
        )
        for pump in result["pumps"]
    ]
    metric_sections[1][2].extend(pump_rows)

    section_colors: dict[int, str] = {}
    row_index = header_row + 1
    formats: dict[int, str] = {}
    for title, color, metrics in metric_sections:
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=last_column)
        section_cell = sheet.cell(row_index, 1, title)
        section_cell.fill = PatternFill("solid", fgColor=color)
        section_cell.font = Font(color="FFFFFF", bold=True, size=11)
        section_cell.alignment = Alignment(vertical="center")
        section_colors[row_index] = color
        sheet.row_dimensions[row_index].height = 22
        row_index += 1

        for key, label, unit, note, value_type in metrics:
            sheet.cell(row_index, 1, label)
            sheet.cell(row_index, 2, unit)
            sheet.cell(row_index, 3, note)
            for year_offset, annual_item in enumerate(result["annual"]):
                column = first_year_column + year_offset
                if key.startswith("@"):
                    value = result["assumptions"].get(key[1:], 0)
                elif key.startswith("#pump:"):
                    nominal = key.split(":", 1)[1]
                    value = annual_item.get("pumpInstallCounts", {}).get(nominal, 0)
                else:
                    value = annual_item.get(key, 0)
                sheet.cell(row_index, column, value)
            formats[row_index] = value_type
            row_index += 1

    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=last_column)
    sheet.cell(row_index, 1, "Контроль расчета").fill = PatternFill("solid", fgColor="7F8C8D")
    sheet.cell(row_index, 1).font = Font(color="FFFFFF", bold=True)
    control_section_row = row_index
    row_index += 1
    controls = [
        (
            "FCF = EBITDA − налог на прибыль − CAPEX",
            [annual["fcfM"] - annual["ebitdaM"] + annual["profitTaxM"] + annual["capexM"] for annual in result["annual"]],
        ),
        (
            "Стоимость остановок/запусков = количество × 1 млн руб.",
            [annual["startStopCostM"] - annual["startStopCount"] for annual in result["annual"]],
        ),
        (
            "Стоимость переводов = количество × 5 млн руб.",
            [annual["conversionOpexM"] - annual["conversionCount"] * 5 for annual in result["annual"]],
        ),
    ]
    for label, values in controls:
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 2, "0 = ОК")
        for year_offset, value in enumerate(values):
            sheet.cell(row_index, first_year_column + year_offset, value)
        formats[row_index] = "control"
        row_index += 1

    warnings = result["diagnostics"].get("warnings", [])
    if warnings:
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=last_column)
        sheet.cell(row_index, 1, "Предупреждения качества данных").fill = PatternFill("solid", fgColor="C55A11")
        sheet.cell(row_index, 1).font = Font(color="FFFFFF", bold=True)
        row_index += 1
        for warning in warnings:
            sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=last_column)
            sheet.cell(row_index, 1, f"• {warning}")
            sheet.cell(row_index, 1).font = Font(color="9C5700")
            sheet.cell(row_index, 1).alignment = Alignment(wrap_text=True)
            row_index += 1

    thin = Side(style="thin", color="D9E2F3")
    for row in range(header_row + 1, row_index):
        if row in section_colors or row == control_section_row:
            continue
        for column in range(1, last_column + 1):
            cell = sheet.cell(row, column)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=(column <= 3))
        if row % 2 == 0:
            for column in range(1, last_column + 1):
                sheet.cell(row, column).fill = PatternFill("solid", fgColor="F7FAFC")
        sheet.cell(row, 1).font = Font(color="17324D")
        sheet.cell(row, 3).font = Font(color="66788A", italic=True, size=9)

    for row, value_type in formats.items():
        number_format = {
            "integer": '#,##0;[Red]-#,##0;–',
            "money": '#,##0.000;[Red]-#,##0.000;–',
            "number": '#,##0.000;[Red]-#,##0.000;–',
            "rubles": '#,##0;[Red]-#,##0;–',
            "percent": '0.0"%"',
            "ratio": '0.000',
            "control": '0.000000;[Red]-0.000000;0.000000',
        }[value_type]
        for column in range(first_year_column, last_column + 1):
            sheet.cell(row, column).number_format = number_format

    value_range = f"{get_column_letter(first_year_column)}{header_row + 1}:{last_letter}{row_index - 1}"
    sheet.conditional_formatting.add(
        value_range,
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FCE8E6")),
    )

    sheet.freeze_panes = "D6"
    sheet.column_dimensions["A"].width = 44
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 58
    for column in range(first_year_column, last_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 16
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.print_area = f"A1:{last_letter}{row_index - 1}"
    sheet.sheet_properties.outlinePr.summaryBelow = True
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(output_path)
    return output_path
